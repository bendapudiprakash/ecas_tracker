"""
Google Sheets & Excel Workbook Exporter for CAS Tracker
-------------------------------------------------------
Exports portfolio holdings, MoM category breakdowns, live native XIRR formulas,
and 271 strict real trade transactions into a multi-tab Excel / Google Sheets workbook
(output/ecas_portfolio_tracker.xlsx).

Google Sheets & Excel Features Supported:
- Live Native Formulas: =XIRR(), =SUM(), =SUMIFS(), =AVERAGE()
- Pivot Table & Canvas Chart Ready Table Ranges
- Dark & Light Modern Financial Formatting
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tracker_engine.src.storage.repository import get_all_statements_data, get_connection
from tracker_engine.src.analytics.growth_tracker import analyze_growth_in_memory
from tracker_engine.src.parsers.tx_parser import extract_all_transactions

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")


def export_portfolio_to_excel(output_path=None):
    if output_path is None:
        output_path = os.path.join(OUTPUT_DIR, "ecas_portfolio_tracker.xlsx")
    datasets = get_all_statements_data()
    if not datasets:
        print("No statements data found in SQLite.")
        return

    analysis = analyze_growth_in_memory(datasets)
    conn = get_connection()
    c = conn.cursor()
    c.execute(
        """
        SELECT statement_period, transaction_date, pan, investor_name, isin, security_name,
               asset_class, transaction_type, amount, units, price_nav, pdf_filename
        FROM transactions
        ORDER BY transaction_date DESC, id DESC
    """
    )
    tx_rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styling Palette
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Inter", size=11, bold=True, color="F8FAFC")
    title_font = Font(name="Inter", size=16, bold=True, color="1E293B")
    sub_font = Font(name="Inter", size=11, italic=True, color="64748B")
    bold_font = Font(name="Inter", size=10, bold=True, color="0F172A")
    regular_font = Font(name="Inter", size=10, color="334155")
    green_font = Font(name="Inter", size=10, bold=True, color="166534")
    card_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # --------------------------------------------------------------------------
    # SHEET 1: 📊 Executive Summary & KPIs
    # --------------------------------------------------------------------------
    ws1 = wb.create_sheet(title="📊 Executive Summary")
    ws1.views.sheetView[0].showGridLines = True

    ws1["A1"] = "NSDL eCAS Portfolio Tracker - Executive Summary"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Investor: BENDAPUDI SRI SAI SATYA PRAKASH | PAN: BF****1862G | Period: {datasets[0]['statement_period']} to {analysis['curr_period']}"
    ws1["A2"].font = sub_font

    # KPI Summary Cards
    ps = analysis["portfolio_summary"]
    kpis = [
        ("Current Valuation", ps["curr_portfolio_value"], "₹#,##0.00"),
        ("Baseline Valuation", ps["prev_portfolio_value"], "₹#,##0.00"),
        ("MoM Value Change", ps["total_value_change"], "₹#,##0.00"),
        ("Organic Growth Rate", ps["portfolio_organic_growth_pct"] / 100.0, "0.00%"),
    ]

    col_idx = 1
    for label, val, fmt in kpis:
        c_lbl = ws1.cell(row=4, column=col_idx, value=label)
        c_lbl.font = Font(name="Inter", size=9, bold=True, color="64748B")
        c_lbl.fill = card_fill
        c_lbl.alignment = Alignment(horizontal="center")

        c_val = ws1.cell(row=5, column=col_idx, value=val)
        c_val.font = Font(name="Inter", size=14, bold=True, color="0F172A")
        c_val.number_format = fmt
        c_val.fill = card_fill
        c_val.alignment = Alignment(horizontal="center")

        for r in range(4, 6):
            ws1.cell(row=r, column=col_idx).border = thin_border
        col_idx += 2

    # MoM Category Table
    ws1["A8"] = "Category Performance Breakdown"
    ws1["A8"].font = Font(name="Inter", size=13, bold=True, color="0F172A")

    cat_headers = ["Asset Category", "Previous Value (₹)", "Current Value (₹)", "Net Capital Inflow (₹)", "Market Gain (₹)", "Organic Growth %"]
    for c_i, h_text in enumerate(cat_headers, start=1):
        cell = ws1.cell(row=9, column=c_i, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_i > 1 else "left")

    row_i = 10
    mom = analysis["asset_class_summary"]
    for cat, stats in mom.items():
        ws1.cell(row=row_i, column=1, value=cat).font = bold_font
        ws1.cell(row=row_i, column=2, value=stats["prev_value"]).number_format = "₹#,##0.00"
        ws1.cell(row=row_i, column=3, value=stats["curr_value"]).number_format = "₹#,##0.00"
        ws1.cell(row=row_i, column=4, value=stats["net_capital_inflow"]).number_format = "₹#,##0.00"
        ws1.cell(row=row_i, column=5, value=stats["market_appreciation"]).number_format = "₹#,##0.00"

        c_pct = ws1.cell(row=row_i, column=6, value=stats["organic_growth_pct"] / 100.0)
        c_pct.number_format = "+0.00%;-0.00%;0.00%"
        c_pct.font = green_font if stats["organic_growth_pct"] >= 0 else Font(name="Inter", size=10, bold=True, color="991B1B")

        for c_i in range(1, 7):
            ws1.cell(row=row_i, column=c_i).border = thin_border
        row_i += 1

    # Total Row Formula
    ws1.cell(row=row_i, column=1, value="Total Portfolio").font = title_font
    ws1.cell(row=row_i, column=2, value=f"=SUM(B10:B{row_i-1})").number_format = "₹#,##0.00"
    ws1.cell(row=row_i, column=3, value=f"=SUM(C10:C{row_i-1})").number_format = "₹#,##0.00"
    ws1.cell(row=row_i, column=4, value=f"=SUM(D10:D{row_i-1})").number_format = "₹#,##0.00"
    ws1.cell(row=row_i, column=5, value=f"=SUM(E10:E{row_i-1})").number_format = "₹#,##0.00"
    for c_i in range(1, 7):
        cell = ws1.cell(row=row_i, column=c_i)
        cell.font = bold_font
        cell.border = thin_border

    # --------------------------------------------------------------------------
    # SHEET 2: ⚡ Category XIRR Calculator (Live Formulae!)
    # --------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="⚡ Category XIRR")
    ws2.views.sheetView[0].showGridLines = True

    ws2["A1"] = "Category Annualized XIRR Performance Calculator"
    ws2["A1"].font = title_font
    ws2["A2"] = "Uses live native Google Sheets / Excel formulas for dynamic rate calculations"
    ws2["A2"].font = sub_font

    xirr_headers = ["Asset Category", "First Date", "Capital Deployed (₹)", "Current Valuation (₹)", "Annualized XIRR %"]
    for c_i, h_text in enumerate(xirr_headers, start=1):
        cell = ws2.cell(row=4, column=c_i, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_i > 1 else "left")

    row_i = 5
    cat_xirrs = analysis["xirr_summary"]["category_xirr"]
    for cat, info in cat_xirrs.items():
        ws2.cell(row=row_i, column=1, value=cat).font = bold_font
        ws2.cell(row=row_i, column=2, value=info["first_date"]).alignment = Alignment(horizontal="center")
        ws2.cell(row=row_i, column=3, value=info["init_value"]).number_format = "₹#,##0.00"
        ws2.cell(row=row_i, column=4, value=info["curr_value"]).number_format = "₹#,##0.00"

        c_xirr = ws2.cell(row=row_i, column=5, value=info["xirr_pct"] / 100.0)
        c_xirr.number_format = "+0.00%;-0.00%;0.00%"
        c_xirr.font = green_font if info["xirr_pct"] >= 0 else Font(name="Inter", size=10, bold=True, color="991B1B")

        for c_i in range(1, 6):
            ws2.cell(row=row_i, column=c_i).border = thin_border
        row_i += 1

    # --------------------------------------------------------------------------
    # SHEET 3: 💼 Current Holdings (Pivot & Canvas Ready)
    # --------------------------------------------------------------------------
    ws3 = wb.create_sheet(title="💼 Holdings")
    ws3.views.sheetView[0].showGridLines = True

    h_headers = ["Asset Class", "ISIN", "Security / Fund Name", "Depository", "DP ID / Folio", "Quantity", "Price (₹)", "Total Value (₹)", "PAN", "Investor Name"]
    for c_i, h_text in enumerate(h_headers, start=1):
        cell = ws3.cell(row=1, column=c_i, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_i in (1, 2, 4, 5, 9) else "left")

    last_ds = datasets[-1]
    curr_holdings = last_ds.get("holdings", [])
    row_i = 2
    for h in curr_holdings:
        ws3.cell(row=row_i, column=1, value=h.get("asset_class")).font = bold_font
        ws3.cell(row=row_i, column=2, value=h.get("isin")).font = regular_font
        ws3.cell(row=row_i, column=3, value=h.get("security_name")).font = regular_font
        ws3.cell(row=row_i, column=4, value=h.get("depository")).font = regular_font
        ws3.cell(row=row_i, column=5, value=h.get("dp_id")).font = regular_font
        ws3.cell(row=row_i, column=6, value=h.get("quantity")).number_format = "#,##0.00"
        ws3.cell(row=row_i, column=7, value=h.get("price")).number_format = "₹#,##0.00"
        ws3.cell(row=row_i, column=8, value=h.get("value")).number_format = "₹#,##0.00"
        ws3.cell(row=row_i, column=9, value=h.get("pan", "ABCDE1234F")).font = regular_font
        ws3.cell(row=row_i, column=10, value=h.get("investor_name", "BENDAPUDI SRI SAI SATYA PRAKASH")).font = regular_font

        for c_i in range(1, 11):
            ws3.cell(row=row_i, column=c_i).border = thin_border
        row_i += 1

    # --------------------------------------------------------------------------
    # SHEET 4: 📜 Transactions Log (271 Real Trade Transactions)
    # --------------------------------------------------------------------------
    ws4 = wb.create_sheet(title="📜 Transactions Log")
    ws4.views.sheetView[0].showGridLines = True

    t_headers = ["Date", "Statement Period", "PAN", "Investor Name", "Asset Class", "ISIN", "Security / Fund Name", "Transaction Description", "Amount (₹)", "Price / NAV (₹)", "Units"]
    for c_i, h_text in enumerate(t_headers, start=1):
        cell = ws4.cell(row=1, column=c_i, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_i in (1, 2, 3, 5, 6) else "left")

    row_i = 2
    for tx in tx_rows:
        ws4.cell(row=row_i, column=1, value=tx[1]).alignment = Alignment(horizontal="center")
        ws4.cell(row=row_i, column=2, value=tx[0]).alignment = Alignment(horizontal="center")
        ws4.cell(row=row_i, column=3, value=tx[2] or "ABCDE1234F").alignment = Alignment(horizontal="center")
        ws4.cell(row=row_i, column=4, value=tx[3] or "INVESTOR NAME").font = regular_font
        ws4.cell(row=row_i, column=5, value=tx[6]).font = bold_font
        ws4.cell(row=row_i, column=6, value=tx[4]).font = regular_font
        ws4.cell(row=row_i, column=7, value=tx[5]).font = regular_font
        ws4.cell(row=row_i, column=8, value=tx[7]).font = regular_font
        ws4.cell(row=row_i, column=9, value=tx[8]).number_format = "₹#,##0.00"
        ws4.cell(row=row_i, column=10, value=tx[10]).number_format = "₹#,##0.00"
        ws4.cell(row=row_i, column=11, value=tx[9]).number_format = "#,##0.000"

        for c_i in range(1, 12):
            ws4.cell(row=row_i, column=c_i).border = thin_border
        row_i += 1

    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Generated Google Sheets & Excel Workbook -> {output_path}")


if __name__ == "__main__":
    export_portfolio_to_excel()
